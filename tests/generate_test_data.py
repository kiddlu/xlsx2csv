#!/usr/bin/env python3
"""
生成所有测试数据 - 包括单元测试和真实场景
Generate all test data - both unit tests and real-world scenarios
"""

from openpyxl import Workbook
from openpyxl.styles import numbers
from datetime import datetime, date, time, timedelta
import os
import random

# ============================================================================
# 单元测试数据 - Unit Test Data
# ============================================================================

def create_date_time_test():
    """Test various date and time formats"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dates and Times"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    
    ws['A2'], ws['B2'] = 'Date 2020-01-01', date(2020, 1, 1)
    ws['A3'], ws['B3'] = 'Date 1900-01-01', date(1900, 1, 1)
    ws['A4'], ws['B4'] = 'Date 2024-12-31', date(2024, 12, 31)
    ws['A5'], ws['B5'] = 'DateTime 2020-06-15 14:30:00', datetime(2020, 6, 15, 14, 30, 0)
    ws['A6'], ws['B6'] = 'Time 12:30:45', time(12, 30, 45)
    ws['A7'], ws['B7'] = 'Time 00:00:00', time(0, 0, 0)
    ws['A8'], ws['B8'] = 'Time 23:59:59', time(23, 59, 59)
    
    ws['B2'].number_format = 'YYYY-MM-DD'
    ws['B3'].number_format = 'YYYY-MM-DD'
    ws['B4'].number_format = 'YYYY-MM-DD'
    ws['B5'].number_format = 'YYYY-MM-DD HH:MM:SS'
    ws['B6'].number_format = 'HH:MM:SS'
    ws['B7'].number_format = 'HH:MM:SS'
    ws['B8'].number_format = 'HH:MM:SS'
    
    wb.save('test_data/date_time.xlsx')
    print("✓ date_time.xlsx")

def create_boolean_test():
    """Test boolean values"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Booleans"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    ws['A2'], ws['B2'] = 'TRUE', True
    ws['A3'], ws['B3'] = 'FALSE', False
    ws['A4'], ws['B4'] = 'Text TRUE', 'TRUE'
    ws['A5'], ws['B5'] = 'Text FALSE', 'FALSE'
    ws['A6'], ws['B6'] = 'Number 1', 1
    ws['A7'], ws['B7'] = 'Number 0', 0
    
    wb.save('test_data/boolean.xlsx')
    print("✓ boolean.xlsx")

def create_percentage_test():
    """Test percentage formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Percentages"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    
    data = [
        ('50%', 0.5, '0%'),
        ('100%', 1.0, '0%'),
        ('0%', 0.0, '0%'),
        ('12.5%', 0.125, '0.0%'),
        ('123.45%', 1.2345, '0.00%'),
        ('-25%', -0.25, '0%'),
    ]
    
    for i, (desc, val, fmt) in enumerate(data, 2):
        ws[f'A{i}'], ws[f'B{i}'] = desc, val
        ws[f'B{i}'].number_format = fmt
    
    wb.save('test_data/percentage.xlsx')
    print("✓ percentage.xlsx")

def create_formula_test():
    """Test formulas (should show calculated values)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"
    
    ws['A1'], ws['B1'], ws['C1'] = 'Description', 'Formula', 'Result'
    ws['A2'], ws['B2'], ws['C2'] = 'Addition', '=2+3', 5
    ws['A3'], ws['B3'], ws['C3'] = 'Multiplication', '=4*5', 20
    ws['A4'], ws['B4'], ws['C4'] = 'Division', '=10/2', 5
    ws['A5'], ws['B5'], ws['C5'] = 'Sum', '=SUM(1,2,3,4,5)', 15
    ws['A6'], ws['B6'], ws['C6'] = 'Concatenation', '="Hello"&" "&"World"', 'Hello World'
    
    wb.save('test_data/formulas.xlsx')
    print("✓ formulas.xlsx")

def create_extreme_numbers_test():
    """Test extreme numeric values"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extreme Numbers"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    
    data = [
        ('Very large positive', 1.7976931348623157e+308),
        ('Very large negative', -1.7976931348623157e+308),
        ('Very small positive', 2.2250738585072014e-308),
        ('Very small negative', -2.2250738585072014e-308),
        ('Tiny positive', 1e-100),
        ('Huge integer', 999999999999999),
        ('Precise decimal', 0.123456789012345),
        ('Pi', 3.141592653589793),
        ('E (Euler)', 2.718281828459045),
    ]
    
    for i, (desc, val) in enumerate(data, 2):
        ws[f'A{i}'], ws[f'B{i}'] = desc, val
    
    wb.save('test_data/extreme_numbers.xlsx')
    print("✓ extreme_numbers.xlsx")

def create_mixed_empty_test():
    """Test mixed content with empty cells"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed Empty"
    
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'Col1', 'Col2', 'Col3', 'Col4'
    ws['A2'], ws['C2'], ws['D2'] = 'A', 'C', 'D'  # B2 empty
    ws['A3'], ws['B3'], ws['D3'] = 1, 2, 4  # C3 empty
    # Row 4 all empty
    ws['A5'], ws['B5'], ws['C5'], ws['D5'] = '', 0, False, 'Text'
    ws['A6'] = 'Last'  # B6, C6, D6 empty
    
    wb.save('test_data/mixed_empty.xlsx')
    print("✓ mixed_empty.xlsx")

def create_unicode_test():
    """Test extended Unicode characters"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Unicode Extended"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    
    data = [
        ('Emoji smile', '😀😃😄😁'),
        ('Emoji various', '🎉🎊🎈🎁'),
        ('Chinese', '中文测试数据'),
        ('Japanese', 'テストデータ'),
        ('Korean', '테스트 데이터'),
        ('Arabic', 'بيانات الاختبار'),
        ('Hebrew', 'נתוני בדיקה'),
        ('Greek', 'δοκιμαστικά δεδομένα'),
        ('Cyrillic', 'тестовые данные'),
        ('Math symbols', '∑∫∂√∞≠≈≤≥'),
        ('Currency symbols', '€¥£₹₽₩¢'),
    ]
    
    for i, (desc, val) in enumerate(data, 2):
        ws[f'A{i}'], ws[f'B{i}'] = desc, val
    
    wb.save('test_data/unicode_extended.xlsx')
    print("✓ unicode_extended.xlsx")

def create_long_strings_test():
    """Test very long strings"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Long Strings"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    
    data = [
        ('Short', 'ABC'),
        ('100 chars', 'A' * 100),
        ('1000 chars', 'B' * 1000),
        ('5000 chars', 'C' * 5000),
        ('With newlines', 'Line1\nLine2\nLine3\nLine4\nLine5'),
        ('With tabs', 'Col1\tCol2\tCol3\tCol4'),
        ('Mixed special', 'Quote"Comma,Tab\tNewline\n'),
    ]
    
    for i, (desc, val) in enumerate(data, 2):
        ws[f'A{i}'], ws[f'B{i}'] = desc, val
    
    wb.save('test_data/long_strings.xlsx')
    print("✓ long_strings.xlsx")

def create_escaping_test():
    """Test CSV escaping scenarios"""
    wb = Workbook()
    ws = wb.active
    ws.title = "CSV Escaping"
    
    ws['A1'], ws['B1'] = 'Description', 'Value'
    
    data = [
        ('Simple comma', 'Hello, World'),
        ('Multiple commas', 'One, Two, Three, Four'),
        ('Double quotes', 'He said "Hello"'),
        ('Quote at start', '"Important"'),
        ('Quote at end', 'Important"'),
        ('Multiple quotes', '"Hello" "World" "Test"'),
        ('Comma and quote', 'Say "Hello, Friend"'),
        ('Leading spaces', '   Leading'),
        ('Trailing spaces', 'Trailing   '),
        ('Both spaces', '  Both  '),
        ('Just spaces', '     '),
    ]
    
    for i, (desc, val) in enumerate(data, 2):
        ws[f'A{i}'], ws[f'B{i}'] = desc, val
    
    wb.save('test_data/escaping.xlsx')
    print("✓ escaping.xlsx")

def create_multisheet_complex_test():
    """Test complex multi-sheet scenarios"""
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Data"
    ws1['A1'], ws1['B1'] = 'Name', 'Value'
    ws1['A2'], ws1['B2'] = 'Alpha', 100
    ws1['A3'], ws1['B3'] = 'Beta', 200
    
    ws2 = wb.create_sheet("Empty")
    
    ws3 = wb.create_sheet("Single")
    ws3['A1'] = 'OnlyOne'
    
    ws4 = wb.create_sheet("Sheet With Spaces")
    ws4['A1'], ws4['B1'] = 'Test', 'Data'
    
    ws5 = wb.create_sheet("Sheet123")
    ws5['A1'] = '123'
    
    wb.save('test_data/multisheet_complex.xlsx')
    print("✓ multisheet_complex.xlsx")

def create_number_formats_test():
    """Test various number format codes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Number Formats"
    
    ws['A1'], ws['B1'] = 'Format', 'Value'
    
    formats = [
        ('General', 12345.6789, 'General'),
        ('0', 12345.6789, '0'),
        ('0.00', 12345.6789, '0.00'),
        ('#,##0', 12345.6789, '#,##0'),
        ('#,##0.00', 12345.6789, '#,##0.00'),
        ('0.00E+00', 12345.6789, '0.00E+00'),
        ('#,##0.00;(#,##0.00)', -1234.56, '#,##0.00;(#,##0.00)'),
    ]
    
    for i, (desc, val, fmt) in enumerate(formats, 2):
        ws[f'A{i}'], ws[f'B{i}'] = desc, val
        ws[f'B{i}'].number_format = fmt
    
    wb.save('test_data/number_formats.xlsx')
    print("✓ number_formats.xlsx")

# ============================================================================
# 真实场景数据 - Real-World Scenario Data
# ============================================================================

def create_stock_data_1107():
    """模拟股票交易数据（多sheet）"""
    wb = Workbook()
    
    # Sheet 1: USDX数据（使用字符串日期）
    ws1 = wb.active
    ws1.title = "USDX"
    ws1.append(["Date", "Open", "High", "Low", "Close", "Volume"])
    
    dates = ["2024-11-07", "2024-11-06", "2024-11-05", "2024-11-04", "2024-11-03"]
    for i, date_str in enumerate(dates[:30] if len(dates) >= 30 else dates * 6):
        open_price = 104.5 + random.uniform(-2, 2)
        close_price = open_price + random.uniform(-1, 1)
        high_price = max(open_price, close_price) + random.uniform(0, 0.5)
        low_price = min(open_price, close_price) - random.uniform(0, 0.5)
        volume = random.randint(1000000, 5000000)
        
        ws1.append([date_str, open_price, high_price, low_price, close_price, volume])
        for col in range(2, 6):
            ws1.cell(i+2, col).number_format = '0.00_ '
    
    # Sheet 2: ETFX数据
    ws2 = wb.create_sheet("ETFX")
    ws2.append(["Symbol", "Price", "Change%", "Volume", "MarketCap"])
    
    etfs = ["TQQQ", "SQQQ", "SPXL", "SPXS", "TNA", "TZA", "SOXL", "SOXS"]
    for etf in etfs:
        ws2.append([etf, random.uniform(10, 200), random.uniform(-15, 15), 
                   random.randint(1000000, 50000000), random.uniform(100, 5000)])
        ws2.cell(ws2.max_row, 2).number_format = '0.00_ '
        ws2.cell(ws2.max_row, 3).number_format = '0.00_ '
    
    # Sheet 3: 详细股票数据（带#VALUE!错误）
    ws3 = wb.create_sheet("Detailed")
    ws3.append(["Code", "TO", "TR", "Chg", "Gap", "Shw", "M2", "TSL", "TD", "Brk", "Res", "Di", "Hi", "Rk", "Up", "VR", "PR", "RE", "PE"])
    
    stocks = [
        ["TSLA", 466.99, 5.10, 14.75, 13.22, 4.33, 13.84, 100.46, 30.41, 2, "1🍀", -0.0, 50.38, 1, 1, 105.88, 288.53, 289.59, 72.92],
        ["NVDA", 346.49, 0.98, 4.07, 2.18, 6.7, 9.68, 8.10, 7.32, 2, "2🍀", "#VALUE!", 49.8, 2, -1, -16.84, 145.61, 146.49, 67.38],
        ["JPM", 58.36, 0.85, 11.54, 6.10, 4.1, 9.55, 33.96, 22.61, 2, "0💎", -0.0, 23.12, 13, 21, 184.47, 247.06, 248, 12.93],
        ["V", 31.35, 0.55, 4.81, 4.44, 5.12, 6.04, 42.51, 16.24, 1, "0🍀", -1.0, 19.2, 24, 13, 45.50, 307.40, 309, 29.08],
    ]
    
    for stock_data in stocks:
        ws3.append(stock_data)
        for col in range(2, 19):
            if isinstance(stock_data[col-1], (int, float)):
                ws3.cell(ws3.max_row, col).number_format = '0.00_ '
    
    # Sheet 4: 科学记数法数据
    ws4 = wb.create_sheet("Scientific")
    ws4.append(["Description", "Value"])
    for desc, val in [("Very small positive", 1.23e-100), ("Very small negative", -4.56e-200),
                      ("Tiny positive", 1e-300), ("Normal float", 123.456789)]:
        ws4.append([desc, val])
        ws4.cell(ws4.max_row, 2).number_format = '0.00_ '
    
    wb.save("test_data/stock_data_1107.xlsx")
    print("✓ stock_data_1107.xlsx")

def create_stock_data_1121():
    """另一天的股票数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trading"
    ws.append(["Symbol", "Open", "Close", "High", "Low", "Volume", "Change%", "Volatility"])
    
    symbols = ["AAPL", "GOOGL", "MSFT", "AMZN", "META", "NFLX", "TSLA", "AMD", "NVDA", "INTC"]
    
    for symbol in symbols:
        open_price = random.uniform(50, 500)
        close_price = open_price * (1 + random.uniform(-0.05, 0.05))
        high_price = max(open_price, close_price) * (1 + random.uniform(0, 0.03))
        low_price = min(open_price, close_price) * (1 - random.uniform(0, 0.03))
        
        ws.append([symbol, open_price, close_price, high_price, low_price,
                  random.randint(5000000, 100000000),
                  ((close_price - open_price) / open_price) * 100,
                  (high_price - low_price) / open_price * 100])
        
        for col in [2, 3, 4, 5, 7, 8]:
            ws.cell(ws.max_row, col).number_format = '0.00_ '
    
    wb.save("test_data/stock_data_1121.xlsx")
    print("✓ stock_data_1121.xlsx")

def create_etfx_template():
    """ETF模板数据 - 测试精度边界情况"""
    wb = Workbook()
    ws = wb.active
    ws.append(["Symbol", "Return_1M", "Return_3M", "Return_6M", "Return_1Y"])
    
    data = [
        ["TQQQ", -1.7215466593796844, 90.5005315773218, -3.7068, 5.1572],
        ["SPXL", -0.5714, 66.6442969389961, -4.6256, 2.8103],
        ["TNA", 1.9481, 83.559958858203686, -10.6163, 2.7404],
        ["FNGU", -2.0450145208131629, 92.3753, -15.9187, 2.4158],
        ["TQQQ", 0.7785, 89.093991817730867, -1.4065, 5.0439],
        ["LABU", 1.4339741569492652, -47.6362, -12.2179, -4.5330],
    ]
    
    for row_data in data:
        ws.append(row_data)
        for col in range(2, 6):
            ws.cell(ws.max_row, col).number_format = '0.00_ '
    
    wb.save("test_data/etfx_template.xlsx")
    print("✓ etfx_template.xlsx")

def create_sector_data():
    """行业数据 - 标准格式测试"""
    wb = Workbook()
    ws = wb.active
    ws.append(["Sector", "Weight%", "Return_YTD", "PE_Ratio", "Dividend_Yield"])
    
    sectors = [
        ["Technology", 28.5, 25.67, 28.4, 1.2],
        ["Healthcare", 13.2, 8.45, 22.1, 1.8],
        ["Financials", 11.8, 15.23, 12.5, 2.5],
        ["Consumer Discretionary", 10.9, 12.34, 19.8, 1.5],
        ["Communication Services", 8.7, 18.92, 15.6, 0.9],
        ["Industrials", 8.4, 9.67, 18.2, 2.1],
        ["Consumer Staples", 6.5, 3.21, 20.5, 2.8],
        ["Energy", 4.2, -5.43, 8.9, 3.5],
        ["Utilities", 2.8, 1.23, 16.7, 3.9],
        ["Real Estate", 2.5, -2.34, 35.2, 4.2],
        ["Materials", 2.5, 7.89, 14.3, 2.3],
    ]
    
    for sector_data in sectors:
        ws.append(sector_data)
        for col in range(2, 6):
            ws.cell(ws.max_row, col).number_format = '#,##0.00'
    
    wb.save("test_data/sector.xlsx")
    print("✓ sector.xlsx")

def create_financial_report():
    """财务报表数据 - 真实的随机数据"""
    wb = Workbook()
    ws = wb.active
    ws.append(["Quarter", "Revenue", "NetIncome", "EPS", "Margin%"])
    
    base_revenue = 50000000
    for quarter in ["Q1 2023", "Q2 2023", "Q3 2023", "Q4 2023", "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024"]:
        revenue = int(base_revenue * (1 + random.uniform(-0.1, 0.15)))
        net_income = int(revenue * random.uniform(0.15, 0.25))
        eps = round(net_income / 1000000 * random.uniform(0.8, 1.2), 2)
        margin = round((net_income / revenue) * 100, 2)
        
        ws.append([quarter, revenue, net_income, eps, margin])
        ws.cell(ws.max_row, 2).number_format = '#,##0'
        ws.cell(ws.max_row, 3).number_format = '#,##0'
        ws.cell(ws.max_row, 4).number_format = '0.00_ '
        ws.cell(ws.max_row, 5).number_format = '0.00_ '
    
    wb.save("test_data/financial_report.xlsx")
    print("✓ financial_report.xlsx")

def create_portfolio_tracking():
    """投资组合跟踪数据"""
    wb = Workbook()
    
    # Sheet 1: 持仓
    ws1 = wb.active
    ws1.title = "Holdings"
    ws1.append(["Symbol", "Shares", "AvgCost", "CurrentPrice", "Value", "Gain/Loss", "Gain%"])
    
    holdings = [
        ["AAPL", 100, 150.25, 175.50, 17550, 2525, 16.79],
        ["GOOGL", 50, 2800.00, 2950.75, 147537.50, 7537.50, 5.38],
        ["TSLA", 75, 220.50, 248.00, 18600, 2062.50, 12.47],
        ["MSFT", 120, 380.00, 405.25, 48630, 3030, 6.64],
    ]
    
    for holding in holdings:
        ws1.append(holding)
        for col in [3, 4, 5, 6, 7]:
            ws1.cell(ws1.max_row, col).number_format = '0.00_ '
    
    # Sheet 2: 交易历史（使用字符串日期）
    ws2 = wb.create_sheet("Transactions")
    ws2.append(["Date", "Type", "Symbol", "Shares", "Price", "Amount", "Commission"])
    
    dates = ["2024-01-01", "2024-01-16", "2024-01-31", "2024-02-15", "2024-03-01",
             "2024-03-16", "2024-03-31", "2024-04-15", "2024-04-30", "2024-05-15",
             "2024-05-30", "2024-06-14", "2024-06-29", "2024-07-14", "2024-07-29",
             "2024-08-13", "2024-08-28", "2024-09-12", "2024-09-27", "2024-10-12"]
    
    for date_str in dates:
        trans_type = random.choice(["BUY", "SELL"])
        symbol = random.choice(["AAPL", "GOOGL", "TSLA", "MSFT"])
        shares = random.randint(10, 100)
        price = random.uniform(100, 500)
        amount = shares * price
        commission = amount * 0.001
        
        ws2.append([date_str, trans_type, symbol, shares, price, amount, commission])
        for col in [5, 6, 7]:
            ws2.cell(ws2.max_row, col).number_format = '0.00_ '
    
    wb.save("test_data/portfolio_tracking.xlsx")
    print("✓ portfolio_tracking.xlsx")

# ============================================================================
# 主函数 - Main
# ============================================================================

def main():
    """生成所有测试数据"""
    random.seed(42)  # 固定随机种子，确保可重现
    os.makedirs('test_data', exist_ok=True)
    
    print("生成测试数据...")
    print("\n=== 单元测试数据 ===")
    create_date_time_test()
    create_boolean_test()
    create_percentage_test()
    create_formula_test()
    create_extreme_numbers_test()
    create_mixed_empty_test()
    create_unicode_test()
    create_long_strings_test()
    create_escaping_test()
    create_multisheet_complex_test()
    create_number_formats_test()
    
    print("\n=== 真实场景数据 ===")
    create_stock_data_1107()
    create_stock_data_1121()
    create_etfx_template()
    create_sector_data()
    create_financial_report()
    create_portfolio_tracking()
    
    print("\n✓ 所有测试数据生成完成！")
    print("  - 单元测试: 11个文件")
    print("  - 真实场景: 6个文件")
    print("  - 总计: 17个测试文件")

if __name__ == '__main__':
    main()
